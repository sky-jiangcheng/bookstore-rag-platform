"""
书单导出API - 使用openpyxl生成标准xlsx文件
网关版本：根据RAG_SERVICE_URL决定转发或本地处理
"""

import io
import json
import logging
import os
from datetime import datetime
from typing import List, Optional
from urllib import error, request

from fastapi import APIRouter, Body, Depends, HTTPException, Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from app.api.auth_management import get_current_active_user
from app.core.logging_config import get_logger
from app.models.auth import User

router = APIRouter()
logger = get_logger(__name__)


class ExportBookItem(BaseModel):
    """导出书籍项"""

    book_id: Optional[int] = None
    title: str
    author: Optional[str] = None
    publisher: Optional[str] = None
    category: Optional[str] = None
    price: Optional[float] = None
    stock: Optional[int] = None
    score: Optional[float] = None
    source: Optional[str] = None
    remark: Optional[str] = None


class ExportBookListRequest(BaseModel):
    """导出书单请求"""

    booklist_name: str = Field(default="书单", description="书单名称")
    books: List[ExportBookItem] = Field(..., min_length=1, description="书籍列表")
    budget: Optional[float] = Field(None, description="预算")
    total_price: Optional[float] = Field(None, description="总价")


def _build_xlsx(request: ExportBookListRequest) -> bytes:
    """用openpyxl构建xlsx并返回字节"""
    wb = Workbook()
    ws = wb.active
    ws.title = request.booklist_name[:31]  # Excel sheet名最长31字符

    # 样式定义
    header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    meta_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    meta_font = Font(name="Microsoft YaHei", bold=True, size=11)
    normal_font = Font(name="Microsoft YaHei", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # 元信息区域
    meta_rows = [
        ("书单名称", request.booklist_name),
        ("书籍数量", str(len(request.books))),
        ("总价格", f"¥{sum(b.price or 0 for b in request.books):.2f}"),
    ]
    if request.budget:
        meta_rows.append(("预算", f"¥{request.budget:.2f}"))
    meta_rows.append(("导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")))

    for idx, (label, value) in enumerate(meta_rows, start=1):
        cell_label = ws.cell(row=idx, column=1, value=label)
        cell_label.font = meta_font
        cell_label.fill = meta_fill
        cell_label.border = thin_border

        cell_value = ws.cell(row=idx, column=2, value=value)
        cell_value.font = normal_font
        cell_value.fill = meta_fill
        cell_value.border = thin_border

    # 空行分隔
    data_start_row = len(meta_rows) + 2

    # 表头
    headers = ["序号", "书名", "作者", "出版社", "分类", "价格", "库存", "相关度", "来源", "备注"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=data_start_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # 数据行
    for row_idx, book in enumerate(request.books, start=1):
        row_num = data_start_row + row_idx
        score = book.score or 0
        score_display = f"{int(score * 100)}%" if score <= 1 else f"{int(score)}%"

        values = [
            row_idx,
            book.title,
            book.author or "",
            book.publisher or "",
            book.category or "",
            book.price or 0,
            book.stock or 0,
            score_display,
            book.source or "",
            book.remark or "",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = normal_font
            cell.border = thin_border
            if col_idx in (1, 6, 7, 8):  # 序号、价格、库存、相关度居中
                cell.alignment = center_align
            if col_idx == 6:  # 价格保留两位小数
                cell.number_format = "¥#,##0.00"

    # 列宽自适应
    col_widths = [6, 30, 15, 20, 12, 10, 8, 10, 10, 20]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # 写入字节
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _forward_to_rag_service(request: ExportBookListRequest, authorization: Optional[str] = None) -> Optional[bytes]:
    """将导出请求转发到RAG服务（如果配置了RAG_SERVICE_URL）"""
    rag_service_url = os.getenv("RAG_SERVICE_URL", "").strip()
    if not rag_service_url:
        return None

    try:
        # 构建转发URL
        base_url = rag_service_url.rstrip("/")
        forward_url = f"{base_url}/book-list/export-excel"

        # 准备请求数据
        payload = json.dumps(request.dict(), ensure_ascii=False).encode("utf-8")
        headers = {
            "Content-Type": "application/json",
            "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }
        if authorization:
            headers["Authorization"] = authorization

        # 发送请求
        req = request.Request(
            forward_url,
            data=payload,
            headers=headers,
            method="POST",
        )

        with request.urlopen(req, timeout=30.0) as response:
            content_type = response.headers.get("Content-Type", "")
            if "spreadsheetml.sheet" in content_type:
                return response.read()
            else:
                logger.warning(f"RAG服务返回非Excel内容类型: {content_type}")
                return None

    except error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        logger.error(f"转发到RAG服务失败，HTTP {exc.code}: {body}")
        return None
    except error.URLError as exc:
        logger.error(f"RAG服务不可达: {exc.reason}")
        return None
    except Exception as e:
        logger.error(f"转发请求异常: {e}")
        return None


@router.post("/api/v1/book-list/export-excel")
async def export_book_list_excel(
    request: ExportBookListRequest = Body(...),
    authorization: Optional[str] = None,
    current_user: User = Depends(get_current_active_user),
):
    """
    导出书单为标准xlsx文件

    网关版本：优先转发到RAG服务，失败时本地生成Excel
    """
    try:
        logger.info(
            "导出书单Excel",
            user_id=current_user.id,
            booklist_name=request.booklist_name,
            book_count=len(request.books),
            has_rag_service=bool(os.getenv("RAG_SERVICE_URL")),
        )

        # 尝试转发到RAG服务
        xlsx_bytes = None
        rag_service_url = os.getenv("RAG_SERVICE_URL", "").strip()
        if rag_service_url:
            xlsx_bytes = _forward_to_rag_service(request, authorization)

        # 如果转发失败或无RAG服务，本地生成
        if not xlsx_bytes:
            xlsx_bytes = _build_xlsx(request)
            source = "local" if not rag_service_url else "local_fallback"
        else:
            source = "rag_service"

        logger.info(f"Excel生成成功，来源: {source}")

        # 生成文件名
        safe_name = "".join(c if c.isalnum() or c in (" ", "-", "_") else "_" for c in request.booklist_name)
        date_str = datetime.now().strftime("%Y%m%d")
        filename = f"{safe_name}_{date_str}.xlsx"

        return Response(
            content=xlsx_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{filename}",
                "Access-Control-Expose-Headers": "Content-Disposition",
            },
        )

    except Exception as e:
        logger.error(f"导出Excel失败: {e}")
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


__all__ = ["router"]
